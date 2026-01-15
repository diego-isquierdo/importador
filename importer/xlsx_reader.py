from __future__ import annotations

import re
import unicodedata
from typing import Any

from openpyxl import load_workbook


def _normalize_header_key(s: str) -> str:
    base = unicodedata.normalize("NFKD", s)
    base = "".join(ch for ch in base if not unicodedata.combining(ch))
    base = base.strip().lower()
    base = re.sub(r"[^a-z0-9]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return base


_HEADER_ALIASES: dict[str, str] = {
    # modelo novo de cancelamento / variações comuns
    "identificador": "identificador",
    "titulo": "Título",
    "t_tulo": "Título",
    "responsavel": "Responsavel",
    "respons_vel": "Responsavel",
    "iniciador": "Iniciador",
    "atividade": "Atividade",
    "tipo_cancelamento": "Tipo_cancelamento",
    "tipo_produto": "Tipo_produto",
    "descricao_motivo": "Descrição_motivo",
    "descri_o_motivo": "Descrição_motivo",
    "modulos": "Modulos",

    # modelo STESTE
    "data_emissao": "Data_emissao",
    "prazo_atv": "Prazo_atv",
    "contato": "Contato_cliente",
    "email": "Email_cliente",

    # variações sem acento/regex para colunas do modelo antigo
    "time_equipe": "Time/Equipe",
    "razao_social": "Razão_social",
    "raz_o_social": "Razão_social",
    "nome_fantasia": "Nome_fantasia",
    "contato_cliente": "Contato_cliente",
    "email_cliente": "Email_cliente",
    "habilitada_em": "Habilitada_em",
    "observacoes_gerais": "Observacoes_gerais",
    "servico_mensal": "Servico_mensal",
    "servico_tecnico": "Servico_tecnico",
    "ativacao_hosting": "Ativação_hosting?",
    "gerar_licenca": "Gerar_licenca?",
    "tem_migracao": "Tem_migração",
    "tem_legal": "Tem_legal?",
}


def read_xlsx_rows(path: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active

    header_cells = list(ws[1])
    raw_headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    headers: list[str] = []
    for h in raw_headers:
        if not h:
            headers.append("")
            continue
        key = _normalize_header_key(h)
        headers.append(_HEADER_ALIASES.get(key, h.strip()))

    rows: list[dict[str, Any]] = []
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [(c.value if c.value is not None else "") for c in row_cells[: len(headers)]]

        # ignorar linhas totalmente vazias
        if all((str(v).strip() == "") for v in values):
            continue

        row: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            row[header] = values[col_idx] if col_idx < len(values) else ""

        row["__row_number__"] = row_idx
        rows.append(row)

    return headers, rows
